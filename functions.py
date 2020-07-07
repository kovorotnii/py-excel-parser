import datetime
from dateparser import parse
import json
from openpyxl import load_workbook, workbook, worksheet

def isAccessible(path, mode='r'):
  """ Check if file is accessbile """
  try:
    f = open(path, mode)
    f.close()
  except IOError:
      return False
  return True

def get_excel_sheets(wb):
  """ Return all possible sheets from excel file """
  return wb.sheetnames

def isJSON(income_str):
  """ Check if income string is JSON """
  if type(income_str) != str:
    return False
  else:
    try: 
      json_resp = json.loads(income_str)
    except ValueError:
      return False
    return True

def parse_excel(wb, current_sheet, col_count, row_count):
  # an array which stores column headers
  headers = []
  # an array contains each row as object 
  values = []
  # processed columns counter
  counter = 0
  # sores row as an object with columns as keys
  inter_dict = {}
  for row in wb[current_sheet].iter_rows(min_row=0, max_col=col_count, max_row=row_count):
    for i, cell in enumerate(row):
      counter += 1
      if (len(headers) == col_count):
        # skip none values
        if headers[i] is None:
          continue

        # try to parse date as string to format
        if ('DATE' in headers[i] or 'TIME' in headers[i]):
          if (cell.value is None):
            inter_dict.update({ headers[i]: None })
          
          # cell value not none
          if (cell.value):
            try:
              if (type(cell.value) == datetime.datetime):
                inter_dict.update({ headers[i]: datetime.datetime.strftime(cell.value, '%Y-%m-%dT%H:%M:%SZ')})
              if (type(cell.value) == datetime.time):
                inter_dict.update({ headers[i]: str(cell.value) }) 
            except TypeError as e:
              print('Parsing date error ', e)
        else:
          if not isJSON(cell.value):
            inter_dict.update({ headers[i]: cell.value })
          else:
            inter_dict.update({headers[i]: json.loads(cell.value) })
      if (len(headers) != col_count):
        headers.append(cell.value)
      if (counter == col_count):
        counter = 0
        if (inter_dict):
          values.append(inter_dict)
          inter_dict = {}
          
  return values
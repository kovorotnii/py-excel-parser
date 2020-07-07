from openpyxl import load_workbook, workbook, worksheet
import json
import functions
import click

@click.command()
@click.option('--excel_path', required=True, help='Path to input - excel file, example: /user/income.xlsm')
# @click.argument('excel_path', required=True, default='data.xlsm', )
@click.option('--json_path', default='default_.json', help='Output path to json, example: /user/output.json')
# @click.argument('json_path', default='default_.json')
def load_excel_file(excel_path, json_path):
  """ Load excel file and invoke parser method """ 
   # Processing default path if output path for json is not defined!
  if json_path == 'default_.json':
    try:
      dot_index = excel_path.index('.')
      json_path = f'{excel_path[:dot_index]}.json'
    except ValueError:
      json_path = 'output.json'
    
  print('Checking output file existence.....')
  # Check if file was created earlier
  is_exist = functions.isAccessible(json_path)
  if is_exist:
    print(f'File {json_path} was created earlier! Set another path or delete existing file')
    quit()
  
  print('Output file not found! Try to load excel file!')
  print('Loading work book! Keep patience!')
  try:
    wb = load_workbook(excel_path, data_only=True)
  except FileNotFoundError:
    print(f'Input excel file {excel_path} not found! Try another path!')
    quit()

  # if excel was found, try to parse it  
  main_execution(wb, excel_path, json_path)


def main_execution(wb, excel_path, json_path):
  # target dict
  target_dict = {}
  # each sheet parsing result will be here
  values = []
  # get a list of sheets
  sheets = functions.get_excel_sheets(wb)

  for current_sheet in sheets:
    row_count = wb[current_sheet].max_row
    col_count = wb[current_sheet].max_column

    print('Parsing sheet: ', current_sheet)

    values = functions.parse_excel(wb, current_sheet, col_count, row_count)

    if not values:
      print("Nothing to assign to target object!")
      quit()
    # Get data from parser, assign to dict
    target_dict.update({ current_sheet: values })
    values = []

  try:
    with open(json_path, 'w') as output:
      data = json.dumps(target_dict, ensure_ascii=False)
      output.write(data)
      output.close()

    print('Parsing completed!')
  except IOError as e:
    print('Writing to file error! ', e)
    quit()

if __name__ == "__main__":
  load_excel_file()